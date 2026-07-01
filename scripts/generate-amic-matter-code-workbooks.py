#!/usr/bin/env python3
import json
import subprocess
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
ARTIFACT_DIR = ROOT / "docs" / "lazycodex" / "evidence" / "matter-desktop" / "artifacts"
CURRENT_JSON = ARTIFACT_DIR / "amic-matter-code-candidates-2026-07-01.json"
CURRENT_XLSX = ARTIFACT_DIR / "amic-current-clients-matter-codes-2026-07-01.xlsx"
REVIEW_XLSX = ARTIFACT_DIR / "matter-code-detail-review-2026-07-01.xlsx"


CORRECTIONS = [
    ("ATU", "ATU Partners", "고객명 정정"),
    ("TAKE Foundation", "B&M Holdings", "고객 통합"),
    ("아론", "", "삭제"),
    ("한흥수 외 6명", "한흥수 외 3명", "고객 수 정정"),
    ("강상도 외 16명", "강상도", "고객명 정정"),
    ("에이치엘엘중앙", "유진이엔티", "고객명 정정"),
    ("SMEJ Holdings", "K Enter Holdings Inc.", "고객명 정정"),
    ("K-PLUS", "바이포엠스튜디오", "인수대상은 K-PLUS로 유지"),
    ("Titan / 오윤록 외 1명", "오윤록 외 2명", "고객 통합"),
    ("고기깡패", "", "삭제"),
    ("고구려푸드", "귀한사람들", "고객 통합"),
    ("부산광역시", "", "삭제"),
]


def read_json(path):
    return json.loads(path.read_text(encoding="utf-8"))


def read_previous_json():
    try:
        raw = subprocess.check_output(
            ["git", "show", f"HEAD:{CURRENT_JSON.relative_to(ROOT)}"],
            cwd=ROOT,
            text=True,
        )
        return json.loads(raw)
    except Exception:
        return {"matters": []}


def autosize(ws):
    for column in ws.columns:
        max_len = 0
        letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[letter].width = max(10, max_len + 2)
    ws.freeze_panes = "A2"


def write_rows(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
    for row in rows:
        values = []
        for header in headers:
            value = row.get(header)
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)
            values.append(value)
        ws.append(values)
    autosize(ws)
    return ws


def matter_key(row):
    return (row.get("client_short_name"), row.get("matter_axis"), row.get("matter_detail_type_korean"))


def build_review_rows(current, previous):
    previous_by_key = {}
    for row in previous.get("matters", []):
        previous_by_key.setdefault(matter_key(row), []).append(row)
    previous_by_client_axis = {}
    for row in previous.get("matters", []):
        previous_by_client_axis.setdefault((row.get("client_short_name"), row.get("matter_axis")), []).append(row)

    review_rows = []
    for row in current["matters"]:
        exact_previous = previous_by_key.get(matter_key(row), [])
        prior = exact_previous[0] if exact_previous else None
        if not prior:
            axis_rows = previous_by_client_axis.get((row.get("client_short_name"), row.get("matter_axis")), [])
            prior = axis_rows[0] if len(axis_rows) == 1 else None
        change_type = "unchanged"
        if not prior:
            change_type = "new_or_split"
        elif prior.get("matter_code") != row.get("matter_code"):
            change_type = "schema_or_detail_changed"
        if row.get("review_required"):
            change_type = f"{change_type}+review_required"
        review_rows.append({
            "고객명": row.get("client_display_name"),
            "이전 Matter Code": prior.get("matter_code") if prior else "",
            "제안 Matter Code": row.get("matter_code"),
            "Axis": row.get("matter_axis"),
            "LIT 하위축": row.get("matter_litigation_axis") or "",
            "세부유형": row.get("matter_detail_type_korean"),
            "고객 지위": row.get("client_case_role") or "",
            "고객 지위 confidence": row.get("client_case_role_confidence") or "",
            "증거": row.get("source_ref"),
            "confidence": row.get("confidence"),
            "review_required": row.get("review_required"),
            "change_type": change_type,
        })
    return review_rows


def main():
    current = read_json(CURRENT_JSON)
    previous = read_previous_json()
    matters = current["matters"]
    clients = current["clients"]
    review_rows = build_review_rows(current, previous)

    counts = {}
    for row in matters:
        key = f"{row['matter_axis']}/{row['matter_litigation_axis']}" if row["matter_axis"] == "LIT" else row["matter_axis"]
        counts[key] = counts.get(key, 0) + 1
    review_count = sum(1 for row in matters if row.get("review_required"))
    generic_count = sum(1 for row in matters if row.get("matter_detail_type_korean") in {"민사사건", "형사사건", "행정사건"})

    wb = Workbook()
    ws = wb.active
    ws.title = "요약"
    summary_rows = [
        ("생성일", current["generated_at"]),
        ("고객 수", current["client_count"]),
        ("Matter Code 수", current["matter_count"]),
        ("LIT/CIV", counts.get("LIT/CIV", 0)),
        ("LIT/CRM", counts.get("LIT/CRM", 0)),
        ("LIT/ADM", counts.get("LIT/ADM", 0)),
        ("Advisory", counts.get("Advisory", 0)),
        ("DEAL", counts.get("DEAL", 0)),
        ("Dispute", counts.get("Dispute", 0)),
        ("검토필요", review_count),
        ("민사사건/형사사건/행정사건", generic_count),
    ]
    for row in summary_rows:
        ws.append(row)
    autosize(ws)

    write_rows(
        wb,
        "고객 목록",
        ["client_id", "client_display_name", "client_short_name", "canonical_display_name", "legal_form", "candidate_type", "source_lanes"],
        clients,
    )
    write_rows(
        wb,
        "Matter Code",
        [
            "matter_id",
            "client_display_name",
            "client_short_name",
            "matter_axis",
            "matter_litigation_axis",
            "matter_detail_type_korean",
            "matter_code",
            "client_case_role",
            "client_case_role_confidence",
            "source_lane",
            "source_ref",
            "confidence",
            "review_required",
            "status",
        ],
        matters,
    )
    write_rows(
        wb,
        "검토필요",
        [
            "matter_id",
            "client_display_name",
            "matter_axis",
            "matter_litigation_axis",
            "matter_detail_type_korean",
            "matter_code",
            "client_case_role",
            "client_case_role_confidence",
            "source_ref",
            "confidence",
            "review_required",
        ],
        [row for row in matters if row.get("review_required")],
    )
    write_rows(
        wb,
        "매크로머신",
        [
            "matter_id",
            "client_display_name",
            "matter_axis",
            "matter_litigation_axis",
            "matter_detail_type_korean",
            "matter_code",
            "client_case_role",
            "client_case_role_confidence",
            "source_ref",
            "confidence",
            "review_required",
        ],
        [row for row in matters if row.get("client_display_name") == "매크로머신"],
    )
    write_rows(
        wb,
        "수정 내역",
        ["before", "after", "note"],
        [{"before": before, "after": after, "note": note} for before, after, note in CORRECTIONS],
    )
    wb.save(CURRENT_XLSX)

    review_wb = Workbook()
    review_wb.active.title = "요약"
    review_wb.active.append(["항목", "값"])
    for row in summary_rows:
      review_wb.active.append(row)
    autosize(review_wb.active)
    write_rows(
        review_wb,
        "전수 리뷰",
        ["고객명", "이전 Matter Code", "제안 Matter Code", "Axis", "LIT 하위축", "세부유형", "고객 지위", "고객 지위 confidence", "증거", "confidence", "review_required", "change_type"],
        review_rows,
    )
    write_rows(
        review_wb,
        "검토필요",
        ["고객명", "이전 Matter Code", "제안 Matter Code", "Axis", "LIT 하위축", "세부유형", "고객 지위", "고객 지위 confidence", "증거", "confidence", "review_required", "change_type"],
        [row for row in review_rows if row["review_required"]],
    )
    write_rows(
        review_wb,
        "매크로머신",
        ["고객명", "이전 Matter Code", "제안 Matter Code", "Axis", "LIT 하위축", "세부유형", "고객 지위", "고객 지위 confidence", "증거", "confidence", "review_required", "change_type"],
        [row for row in review_rows if row["고객명"] == "매크로머신"],
    )
    review_wb.save(REVIEW_XLSX)

    print(json.dumps({
        "current_xlsx": str(CURRENT_XLSX),
        "review_xlsx": str(REVIEW_XLSX),
        "clients": len(clients),
        "matters": len(matters),
        "counts": counts,
        "review_required": review_count,
        "generic_count": generic_count,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
